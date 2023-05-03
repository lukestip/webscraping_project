from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font
import keys 
from twilio.rest import Client

'''Find a 'scrappable' cryptocurrencies website where you can scrape the top 5 cryptocurrencies. 
Scrape the information and create a well formatted  Excel spreadsheet report. 
Be creative with your report (as in colors, spacing, currency, text size, etc). 
The report should display the name of the currency, the symbol (if applicable), 
the current price and % change in the last 24 hrs and corresponding price (based on % change)

Furthermore, for Bitcoin and Ethereum, the program should alert you via text if the value 
increases or decrease within $5 of its current value.
Submit your GitHub URL which should contain all the completed files worked in class as well as the above.

NOTE: Remember to ignore your virtual environment in your repository. 
For your keys file, please remove your personal api key information and leave it blank. 
Twilio will automatically detect if it finds your api keys in Github and invalidate it. 
We do not need your keys since we can use our own to run it.'''

#Beautiful Soup
webpage = 'https://coinranking.com/'
page = urlopen(webpage)			
soup = BeautifulSoup(page, 'html.parser')
title = soup.title
print(title.text)

#Twilio
client = Client(keys.account_sid, keys.auth_token)
TwilioNumber = "+15074167105"
myNumber = "+19724397045"

#Excel
wb = xl.Workbook()
ws = wb.active
ws.title = 'Crypto Rankings'
Font1 = Font(name='Arial',size=12,italic=False, bold=True)

ws['A1'] = 'Rank'
ws['A1'].font = Font1
ws['B1'] = 'Title'
ws['B1'].font = Font1
ws['C1'] = 'Symbol'
ws['C1'].font = Font1
ws['D1'] = 'Price'
ws['D1'].font = Font1
ws['E1'] = '%Change (24h)'
ws['E1'].font = Font1
ws['F1'] = '%Change Price'
ws['F1'].font = Font1

ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 16
ws.column_dimensions['F'].width = 16

#Webscraping website
cryptoRows = soup.findAll("tr")
td = cryptoRows[1].findAll('div')

#Top 5 Cryptocurrencies
for x in range (4,9):
    td = cryptoRows[x].findAll('div')

    info = td[0].text.split()
    rank = info[0]
    title = info[1]
    symbol = info[2]
    
    # current price and % change in the last 24 hrs and corresponding price (based on % change)
    priceInfo = td[1].text.split()
    price = float(priceInfo[1].replace(',',''))

    change = float(td[3].text.replace('\n','').replace('%','').replace(' ',''))
    
    changePrice = round(price*change,3)

    #Write in information to sheet
    ws['A'+str(x-2)] = rank
    ws['B'+str(x-2)] = title
    ws['C'+str(x-2)] = symbol
    ws['D'+str(x-2)] = float(price)
    ws['E'+str(x-2)] = str(change)+'%'
    ws['F'+str(x-2)] = float(changePrice)

    #Twilio: for Bitcoin and Ethereum, alert you via text if the value increases or decrease within $5 of its current value.
    if title == 'Bitcoin':
        if abs(changePrice) > 5:
            message = 'ALERT: '+title+' has changed price by more than $5 in the past day!'
            textmsg = client.messages.create(to=myNumber, from_=TwilioNumber, body=message)
            print(textmsg.status)
    elif title == 'Ethereum':
        if abs(changePrice) > 5:
            message = 'ALERT: '+title+' has changed price by more than $5 in the past day!'
            textmsg = client.messages.create(to=myNumber, from_=TwilioNumber, body=message)
            print(textmsg.status)

# Formatting for prices
for cell in ws['D:D']:
    cell.number_format = u'#$ "#,##0.00'

for cell in ws['F:F']:
    cell.number_format = u'#$ "#,##0.00'

wb.save('Webscraping_Project.xlsx')